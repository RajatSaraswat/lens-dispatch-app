import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# === File paths ===
orders_file = "GKB Dispatch Details 19.07.2025(third route).xls.xlsx"
address_file = "EYE GEAR ADDRESS.xlsx"
output_file = "Bluedart_Upload_Ready.xlsx"

# === Load data ===
orders_df = pd.read_excel(orders_file)
address_df = pd.read_excel(address_file)

orders_df.columns = orders_df.columns.str.strip()
address_df.columns = address_df.columns.str.strip()

address_df = address_df.rename(columns={"BranchID": "Store_Code", "BranchAddress": "Store_Address"})
address_df["Store_Code"] = address_df["Store_Code"].astype(str).str.strip()

def extract_store_code(ref):
    if isinstance(ref, str) and "/" in ref:
        return ref.split("/")[0].strip()
    return None

orders_df["Store_Code"] = orders_df["Customer Reference"].apply(extract_store_code)
orders_df["Store_Code"] = orders_df["Store_Code"].astype(str).str.strip()

bluedart_df = address_df[address_df["Delivery By"].str.lower() == "bluedart"]
merged_df = pd.merge(orders_df, bluedart_df, on="Store_Code", how="inner")

today = datetime.today()
today_ddmm = today.strftime("%d%m")
pickup_date = today.strftime("%d-%m-%Y")

# Fixed values
billing_area = "NDA"
billing_customer_code = "991712"
pickup_time = "1930"
shipper_name = "GKB EYECARE"
pickup_address = "B 48 GKB EYECARE"
pickup_pincode = "201305"
product_code = "D"
product_type = "NDOX"
declared_value = 500
register_pickup = "TRUE"
sender = "GKB EYECARE"
receiver_name = "Ben Franklin / EyeGear"
office_closure_time = "1930"

records = []
grouped = merged_df.groupby("Store_Address")

for address, group in grouped:
    store_codes = sorted(group["Store_Code"].unique())
    store_str = "_".join(store_codes)
    pincode = group["Pincode"].iloc[0]
    mobile = group["Mobile"].iloc[0] if pd.notnull(group["Mobile"].iloc[0]) else "9548993667"

    ref_no = f"Ben-{store_codes[0]}-{today_ddmm}" if len(store_codes) == 1 else f"Bulk-{store_str}-{today_ddmm}"

    record = {
        "Reference No": ref_no,
        "Billing Area": billing_area,
        "Billing Customer Code": billing_customer_code,
        "Pickup Date": pickup_date,
        "Pickup Time": pickup_time,
        "Shipper Name": shipper_name,
        "Pickup address": pickup_address,
        "Pickup pincode": pickup_pincode,
        "Company Name": "ben franklin",
        "Delivery address": address,
        "Delivery Pincode": pincode,
        "Product Code": product_code,
        "Product Type": product_type,
        "Pack Type": "",
        "Piece Count": 1,
        "Actual Weight": 0.5,
        "Declared Value": declared_value,
        "Register Pickup": register_pickup,
        "Length": "", "Breadth": "", "Height": "",
        "To Pay Customer": "",
        "Sender": sender,
        "Sender mobile": "",
        "Receiver Telephone": "",
        "Receiver mobile": mobile,
        "Receiver Name": receiver_name,
        "Special Instruction": "",
        "Commodity Detail 1": "", "Commodity Detail 2": "",
        "Commodity Detail 3": "",
        "Reference No 2": "", "Reference No 3": "",
        "OTP Based Delivery": "", "Office Closure time": office_closure_time,
        "AWB No": "", "Status": "", "Message": "", "Cluster Code": "",
        "Destination Area": "", "Destination Location": "", "Pick Up Token No": "",
        "Response pick up date": "", "Transaction Amount": "", "Wallet Balance": "",
        "Available Booking Amount": ""
    }

    records.append(record)

# Define final columns
bluedart_columns = [
    "Reference No", "Billing Area", "Billing Customer Code", "Pickup Date", "Pickup Time",
    "Shipper Name", "Pickup address", "Pickup pincode", "Company Name", "Delivery address",
    "Delivery Pincode", "Product Code", "Product Type", "Pack Type", "Piece Count",
    "Actual Weight", "Declared Value", "Register Pickup", "Length", "Breadth", "Height",
    "To Pay Customer", "Sender", "Sender mobile", "Receiver Telephone", "Receiver mobile",
    "Receiver Name", "Special Instruction", "Commodity Detail 1", "Commodity Detail 2",
    "Commodity Detail 3", "Reference No 2", "Reference No 3", "OTP Based Delivery",
    "Office Closure time", "AWB No", "Status", "Message", "Cluster Code", "Destination Area",
    "Destination Location", "Pick Up Token No", "Response pick up date", "Transaction Amount",
    "Wallet Balance", "Available Booking Amount"
]

# Save to Excel
df = pd.DataFrame(records, columns=bluedart_columns)
df.to_excel(output_file, index=False)

# === Apply color formatting with openpyxl ===
wb = load_workbook(output_file)
ws = wb.active

# Example fills based on your screenshot colors
fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Column A
fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")   # Columns J, K, Z
fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Column N

# Apply colors to header row
for col in range(1, ws.max_column + 1):
    header_cell = ws.cell(row=1, column=col)
    header_cell.font = Font(bold=True)

    col_letter = ws.cell(row=1, column=col).column_letter

    if col_letter == "A":
        header_cell.fill = fill_orange
    elif col_letter in ["J", "K", "Z"]:
        header_cell.fill = fill_green
    elif col_letter == "N":
        header_cell.fill = fill_yellow

# Save with formatting
wb.save(output_file)
print(f"✅ File with formatting saved: {output_file}")
